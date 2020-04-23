import json
import re

from openpyxl import load_workbook, Workbook

FIRST_ROW = 3
WEEK_COL = 'B'
ADDRESS_COL = 'C'
HOUSEHOLD_COL = 'D'
NAME_COL = 'E'
CONTACT_COL = 'F'
NEED_UPDATE_COL = 'G'

def load_address_with_postal_key(): 
    """
    Data set gotten from https://github.com/xkjyeah/singapore-postal-codes
    """
    db = open('buildings.json.data')
    db = json.load(db)
    return { x['POSTAL']: x for x in db}

def load_excel():
    finished = False
    cur_row = FIRST_ROW
    cur_row_str = str(cur_row)
    families = []

    wb = load_workbook("voucher_list.xlsx")
    sheet = wb['Batch 2 ']

    while not finished:
        new_family = {
            'week': sheet[WEEK_COL + cur_row_str].value.strip(),
            'address': sheet[ADDRESS_COL + cur_row_str].value.strip(),
            'household': sheet[HOUSEHOLD_COL + cur_row_str].value,
            'name': sheet[NAME_COL + cur_row_str].value.strip(),
            'contact': sheet[CONTACT_COL + cur_row_str].value,
            'need_update': sheet[NEED_UPDATE_COL + cur_row_str].value,
        }
        families.append(new_family)
        cur_row += 1
        cur_row_str = str(cur_row)
        if not sheet[NAME_COL + cur_row_str].value:
            finished = True
    return families

def wrapper_add_address_map(address_map):
    def add_address(family):
        addresses = address_map
        if family['need_update']:
            address = family['address']
            postal_code = re.findall('\d{6}', address)[0]
            print(postal_code)
            unit_number = re.findall('#\d+-\d+', address)[0]
            family['address'] = addresses[postal_code]['BLK_NO'] + ' ' + addresses[postal_code]['ROAD_NAME'] + ', ' + unit_number + ', (S)' + postal_code
        return family
    return add_address

def write_to_excel(families):
    wb = Workbook()
    sheet = wb.active

    sheet['B2'] = 'weeks'
    sheet['C2'] = 'Address'
    sheet['D2'] = 'No. pax/ household'
    sheet['E2'] = 'Name'
    sheet['F2'] = 'Contact'
    sheet['G2'] = ''

    cur_row = 3
    cur_row_str = str(cur_row)

    for family in families:
        sheet[WEEK_COL + cur_row_str] = family['week']
        sheet[ADDRESS_COL + cur_row_str] = family['address']
        sheet[HOUSEHOLD_COL + cur_row_str] = family['household']
        sheet[NAME_COL + cur_row_str] = family['name']
        sheet[CONTACT_COL + cur_row_str] = family['contact']
        sheet[NEED_UPDATE_COL + cur_row_str] = family['need_update']
        cur_row += 1
        cur_row_str = str(cur_row)
    wb.save("new.xlsx")

if __name__ == "__main__":
    address_map = load_address_with_postal_key()
    families = load_excel()
    updated_familes = map(wrapper_add_address_map(address_map), families)
    write_to_excel(updated_familes)

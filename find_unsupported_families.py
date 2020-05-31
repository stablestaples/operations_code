import json
import re

from openpyxl import load_workbook, Workbook

SUPPORTED_FAMILIES_FILENAME = "needs/sent.xlsx"
SUPPORTED_FAMILIES_SHEET_NAME = "Check Items to Purchase"
SUPPORTED_FAMILES_ID_COL = "F"
SUPPORTED_FAMILES_NAME_COL = "G"
SUPPORTED_FAMILES_ITEM_COL = "H"
SUPPORTED_FAMILES_FIRST_ROW = 2

MASTER_SHEET_ONE_FILENAME = "needs/needs.xlsx"
MASTER_SHEET_ONE_SHEETNAME = "Sheet1"
MASTER_SHEET_ONE_NAME_COL = "E"
MASTER_SHEET_ONE_REMARKS_COL = "O"
MASTER_SHEET_ONE_FIRST_ROW = 3

MASTER_SHEET_THREE_FILENAME = "needs/needs.xlsx"
MASTER_SHEET_THREE_SHEETNAME = "Sheet3"
MASTER_SHEET_THREE_NAME_COL = "C"
MASTER_SHEET_THREE_REMARKS_COL = "M"
MASTER_SHEET_THREE_FIRST_ROW = 3

MASTER_SHEET_FOUR_FILENAME = "needs/needs.xlsx"
MASTER_SHEET_FOUR_SHEETNAME = "Sheet4"
MASTER_SHEET_FOUR_NAME_COL = "C"
MASTER_SHEET_FOUR_REMARKS_COL = "M"
MASTER_SHEET_FOUR_FIRST_ROW = 2

MASTER_SHEET_FIVE_FILENAME = "needs/needs.xlsx"
MASTER_SHEET_FIVE_SHEETNAME = "Sheet5"
MASTER_SHEET_FIVE_NAME_COL = "C"
MASTER_SHEET_FIVE_REMARKS_COL = "M"
MASTER_SHEET_FIVE_FIRST_ROW = 2

MASTER_SHEET_SIX_FILENAME = "needs/needs.xlsx"
MASTER_SHEET_SIX_SHEETNAME = "Sheet6"
MASTER_SHEET_SIX_NAME_COL = "C"
MASTER_SHEET_SIX_REMARKS_COL = "M"
MASTER_SHEET_SIX_FIRST_ROW = 2


def load_supported_excel():
    finished = False
    cur_row = SUPPORTED_FAMILES_FIRST_ROW
    cur_row_str = str(cur_row)
    families = {}
    name_to_family_reverse_lookup = {}

    wb = load_workbook(SUPPORTED_FAMILIES_FILENAME)
    sheet = wb[SUPPORTED_FAMILIES_SHEET_NAME]

    while not finished:
        if sheet[SUPPORTED_FAMILES_ID_COL + cur_row_str].value not in families:
            families[sheet[SUPPORTED_FAMILES_ID_COL + cur_row_str].value] = {
                'supplies': []
            }
        if sheet[SUPPORTED_FAMILES_NAME_COL + cur_row_str].value not in name_to_family_reverse_lookup:
            name_to_family_reverse_lookup[sheet[SUPPORTED_FAMILES_NAME_COL + cur_row_str].value] = sheet[SUPPORTED_FAMILES_ID_COL + cur_row_str].value
        families[sheet[SUPPORTED_FAMILES_ID_COL + cur_row_str].value]['supplies'].append((sheet[SUPPORTED_FAMILES_NAME_COL + cur_row_str].value, sheet[SUPPORTED_FAMILES_ITEM_COL + cur_row_str].value))

        cur_row += 1
        cur_row_str = str(cur_row)
        if not sheet[SUPPORTED_FAMILES_ID_COL + cur_row_str].value:
            finished = True
    return families,name_to_family_reverse_lookup

def update_needs_excel_with_support_data(supported_families, reverse_lookup, filename, sheetname, name_col, remarks_col, first_row):
    finished = False
    cur_row = first_row
    cur_row_str = str(cur_row)

    wb = load_workbook(filename)
    sheet = wb[sheetname]

    while not finished:
        if sheet[name_col + cur_row_str].value in reverse_lookup:
            sheet[remarks_col+cur_row_str].value = reverse_lookup[sheet[name_col + cur_row_str].value] + " " + str(supported_families[reverse_lookup[sheet[name_col + cur_row_str].value]]["supplies"])
        cur_row += 1
        cur_row_str = str(cur_row)
        if not sheet[name_col + cur_row_str].value:
            finished = True
    wb.save(filename)




if __name__ == "__main__":
    supported_families, reverse_lookup = load_supported_excel()
    update_needs_excel_with_support_data(supported_families, reverse_lookup, MASTER_SHEET_ONE_FILENAME, MASTER_SHEET_ONE_SHEETNAME, MASTER_SHEET_ONE_NAME_COL, MASTER_SHEET_ONE_REMARKS_COL, MASTER_SHEET_ONE_FIRST_ROW)
    update_needs_excel_with_support_data(supported_families, reverse_lookup, MASTER_SHEET_THREE_FILENAME, MASTER_SHEET_THREE_SHEETNAME, MASTER_SHEET_THREE_NAME_COL, MASTER_SHEET_THREE_REMARKS_COL, MASTER_SHEET_THREE_FIRST_ROW)
    update_needs_excel_with_support_data(supported_families, reverse_lookup, MASTER_SHEET_FOUR_FILENAME, MASTER_SHEET_FOUR_SHEETNAME, MASTER_SHEET_FOUR_NAME_COL, MASTER_SHEET_FOUR_REMARKS_COL, MASTER_SHEET_FOUR_FIRST_ROW)
    update_needs_excel_with_support_data(supported_families, reverse_lookup, MASTER_SHEET_FIVE_FILENAME, MASTER_SHEET_FIVE_SHEETNAME, MASTER_SHEET_FIVE_NAME_COL, MASTER_SHEET_FIVE_REMARKS_COL, MASTER_SHEET_FIVE_FIRST_ROW)
    update_needs_excel_with_support_data(supported_families, reverse_lookup, MASTER_SHEET_SIX_FILENAME, MASTER_SHEET_SIX_SHEETNAME, MASTER_SHEET_SIX_NAME_COL, MASTER_SHEET_SIX_REMARKS_COL, MASTER_SHEET_SIX_FIRST_ROW)
    
